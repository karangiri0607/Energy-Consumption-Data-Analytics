import openpyxl
excel_file =['C:\\Local Disk D\\Karan\\14_June.xlsx',]
for file in excel_file:
    wb = openpyxl.load_workbook(file)
    worksheet =wb["Sheet1"]
    worksheet['E']
    var=worksheet.cell(row=6,column=5).value
    print(worksheet.cell(row=6,column=5).value)
    worksheet['E']
    var1=worksheet.cell(row=7,column=5).value
    print(worksheet.cell(row=7,column=5).value)
    worksheet['D']
    var2=worksheet.cell(row=10,column=4).value
    print(worksheet.cell(row=10,column=4).value)
    worksheet['D']
    var3=worksheet.cell(row=11,column=4).value
    print(worksheet.cell(row=11,column=4).value)
    worksheet['C']
    var4=worksheet.cell(row=12,column=3).value
    print(worksheet.cell(row=12,column=3).value)
    worksheet['D']
    var5=worksheet.cell(row=12,column=4).value
    print(worksheet.cell(row=12,column=4).value)
    worksheet['E']
    var6=worksheet.cell(row=12,column=5).value
    print(worksheet.cell(row=12,column=5).value)
    worksheet['D']
    var7=worksheet.cell(row=15,column=4).value
    print(worksheet.cell(row=15,column=4).value)
    worksheet['D']
    var8=worksheet.cell(row=16,column=4).value
    print(worksheet.cell(row=16,column=4).value)
    worksheet['D']
    var9=worksheet.cell(row=17,column=4).value
    print(worksheet.cell(row=17,column=4).value)
    worksheet['D']
    var10=worksheet.cell(row=18,column=4).value
    print(worksheet.cell(row=18,column=4).value)
    worksheet['D']
    var11=worksheet.cell(row=19,column=4).value
    print(worksheet.cell(row=19,column=4).value)
    worksheet['D']
    var12=worksheet.cell(row=20,column=4).value
    print(worksheet.cell(row=20,column=4).value)
    worksheet['D']
    var13=worksheet.cell(row=21,column=4).value
    print(worksheet.cell(row=21,column=4).value)
    worksheet['D']
    var14=worksheet.cell(row=22,column=4).value
    print(worksheet.cell(row=22,column=4).value)
    worksheet['E']
    var15=worksheet.cell(row=22,column=5).value
    print(worksheet.cell(row=22,column=5).value)
    worksheet['C']
    var16=worksheet.cell(row=25,column=3).value
    print(worksheet.cell(row=25,column=3).value)
    worksheet['D']
    var17=worksheet.cell(row=25,column=4).value
    print(worksheet.cell(row=25,column=4).value)
    worksheet['D']
    var18=worksheet.cell(row=26,column=4).value
    print(worksheet.cell(row=26,column=4).value)
    worksheet['E']
    var19=worksheet.cell(row=26,column=5).value
    print(worksheet.cell(row=26,column=5).value)
    worksheet['C']
    var20=worksheet.cell(row=29,column=3).value
    print(worksheet.cell(row=29,column=3).value)
    worksheet['D']
    var21=worksheet.cell(row=29,column=4).value
    print(worksheet.cell(row=29,column=4).value)
    worksheet['D']
    var22=worksheet.cell(row=30,column=4).value
    print(worksheet.cell(row=30,column=4).value)
    worksheet['D']
    var23=worksheet.cell(row=31,column=4).value
    print(worksheet.cell(row=31,column=4).value)
    worksheet['E']
    var24=worksheet.cell(row=31,column=5).value
    print(worksheet.cell(row=31,column=5).value)
    

excel_file1 =['C:\\Local Disk D\\Karan\\Energy-Consumption-Data Copy1.xlsx',]
for file in excel_file1:
    wb = openpyxl.load_workbook(file)
    worksheet =wb["Sheet1"]
    worksheet['B23'] =var
    worksheet =wb["Sheet1"]
    worksheet['C23'] =var1
    worksheet =wb["Sheet1"]
    worksheet['E23'] =var2
    worksheet =wb["Sheet1"]
    worksheet['F23'] =var3
    worksheet =wb["Sheet1"]
    worksheet['G23'] =var4
    worksheet =wb["Sheet1"]
    worksheet['H23'] =var5
    worksheet =wb["Sheet1"]
    worksheet['D23'] =var6
    worksheet =wb["Sheet1"]
    worksheet['J23'] =var7
    worksheet =wb["Sheet1"]
    worksheet['K23'] =var8
    worksheet =wb["Sheet1"]
    worksheet['L23'] =var9
    worksheet =wb["Sheet1"]
    worksheet['M23'] =var10
    worksheet =wb["Sheet1"]
    worksheet['N23'] =var11
    worksheet =wb["Sheet1"]
    worksheet['O23'] =var12
    worksheet =wb["Sheet1"]
    worksheet['P23'] =var13
    worksheet =wb["Sheet1"]
    worksheet['Q23'] =var14
    worksheet =wb["Sheet1"]
    worksheet['I23'] =var15
    worksheet =wb["Sheet1"]
    worksheet['S23'] =var16
    worksheet =wb["Sheet1"]
    worksheet['T23'] =var17
    worksheet =wb["Sheet1"]
    worksheet['U23'] =var18
    worksheet =wb["Sheet1"]
    worksheet['R23'] =var19
    worksheet =wb["Sheet1"]
    worksheet['W23'] =var20
    worksheet =wb["Sheet1"]
    worksheet['X23'] =var21
    worksheet =wb["Sheet1"]
    worksheet['Y23'] =var22
    worksheet =wb["Sheet1"]
    worksheet['Z23'] =var23
    worksheet =wb["Sheet1"]
    worksheet['V23'] =var24
    wb.save(file)